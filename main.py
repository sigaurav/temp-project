from openpyxl import load_workbook
import pandas as pd
import tkinter.ttk as ttk
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter
import os
import support as sp


class SegregateFiles:
    managers_lst = []
    blank_cols = 0
    formula_column = {}
    def process_masterfile(filepath: str, output_path: str, progressbar: ttk.Progressbar):
        """:param filepath: Path of input file
        :param output_path: output directory"""

        file_name = os.path.basename(filepath)
        wb = load_workbook(filepath)
        sheets = wb.sheetnames
        sheet_name = wb[sheets[0]]
        master_df = pd.DataFrame(sheet_name.values)
        master_df.columns = master_df.iloc[0]
        master_df = master_df[1:]
        master_df.reset_index()

        header_style = sp.capture_format(sheet_name)
        managers_lst = master_df.Manager.unique().tolist()
        filenum = 0
        totalManagers = len(managers_lst)
        progressbar['maximum'] = 100
        SegregateFiles.formula_column = SegregateFiles.get_formula_columns(master_df)
        for manager in managers_lst:
            temp_df = master_df[master_df['Manager'] == manager]
            temp_df.reset_index(inplace=True, drop=True)
            SegregateFiles.update_formula(temp_df, SegregateFiles.formula_column)
            op_file_name = manager + "_" + file_name
            sp.export_file(temp_df, output_path + "/" + op_file_name, header_style)
            filenum = filenum+1
            progressbar['value'] = round((filenum/totalManagers)*100, 2)
            progressbar.update()


    def update_formula(df: pd.DataFrame, formula_col: str):
        '''
        Function to update formulas for each Manager
        :param df:  DataFrame for one specific manager.
        '''
        for _col in formula_col:
            col_alpha = formula_col[_col][0]
            formula = formula_col[_col][1]
            index = 2
            for ind, row in df.iterrows():
                df.at[ind, _col] = Translator(formula, origin=col_alpha + '2').translate_formula(col_alpha + str(index))
                index = index + 1

    def get_formula_columns(df: pd.DataFrame):
        '''
        Function to find formulae based columns and rename blank columns
        @param: df master df
        '''

        formula_cols = {}
        for i in range(len(df.columns.tolist())):
            col = df.columns[i]
            if pd.isna(col) or str(col).strip() == '':
                df.rename(columns={col: 'BlankCol' + str(SegregateFiles.blank_cols)}, inplace=True)
                col = 'BlankCol' + str(SegregateFiles.blank_cols)
                SegregateFiles.blank_cols = SegregateFiles.blank_cols + 1

            col_value = str(df[df.columns[i]].values[0])
            if col_value.startswith('='):
                if col not in formula_cols:
                    formula_cols[col] = [get_column_letter(i + 1), col_value]
                else:
                    df.rename(columns={col: col + "1"}, inplace=True)
                    formula_cols[col + "1"] = [get_column_letter(i + 1), col_value]

        return formula_cols


class MergeFiles:
    df_columns = []
    formula_column = {}
    file_num = 1
    def merge_files(file_list: tuple, file_dir: str, progressbar: ttk.Progressbar):

        master_df = pd.DataFrame()
        # Read all excel files
        header_file = {}
        total_files = len(file_list)
        progressbar['maximum'] = 100
        for file in file_list:
            wb = load_workbook(file)
            sheet_name = wb[wb.sheetnames[0]]
            temp_df = pd.DataFrame(sheet_name.values)
            if file_num == 1:
                MergeFiles.df_columns = temp_df.iloc[0]
                header_file = sp.capture_format(sheet_name)

            temp_df = temp_df[1:]
            temp_df.columns = MergeFiles.df_columns
            master_df = pd.concat([master_df, temp_df], axis=0)
            file_num = file_num + 1
            progressbar['value'] = round((file_num/total_files) * 100, 2)
            progressbar.update()

        MergeFiles.formula_column = SegregateFiles.get_formula_columns(master_df)
        master_df.reset_index(inplace=True, drop=True)
        SegregateFiles.update_formula(master_df, MergeFiles.formula_column)
        sp.export_file(master_df, file_dir + "/" + "MergedOutput.xlsx", header_file)


        print(f'Merged {file_num} files')
