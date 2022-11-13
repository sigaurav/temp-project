import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy

def export_file(manager_df: pd.DataFrame, output_file: str, header_style: dict):
    manager_df.to_csv(output_file, index=False)

    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
    manager_df.to_excel(writer, sheet_name='Sheet1', index=False)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    cell_format_header = workbook.add_format({'bold': True, 'font_color': 'white', 'font_size': 14,
                                              'bg_color': '#3b5998', 'align': 'center', 'valign': 'vcenter'})

    for col_num, value in enumerate(manager_df.columns.values):
        worksheet.write(0, col_num, value, cell_format_header)
    writer.close()
    apply_format(output_file, header_style)


def apply_format(output_file: str, header_style: dict):
    """
    Function to apply format to headers while merging and splitting files
    :return:
    """
    wb = load_workbook(output_file)
    sheets = wb.sheetnames
    sheet_name = wb[sheets[0]]
    for row in sheet_name.rows:
        for cell in row:
            cell.font = copy(header_style[cell.value][0])
            cell.border = copy(header_style[cell.value][1])
            cell.fill = copy(header_style[cell.value][2])
            cell.number_format = copy(header_style[cell.value][3])
            cell.alignment = copy(header_style[cell.value][4])
        break;

    wb.save(output_file)
    wb.close()

def capture_format(sh: Worksheet):
    """
    Function to capture header format of input file
    :return:
    """
    blank_cols = 0

    cell_format = {}
    for row in sh.rows:
        for cell in row:
            if str(cell.value).strip() == '' or cell.value is None:
                cell.value= 'BlankCol' + str(blank_cols)
                blank_cols = blank_cols + 1
            if cell.has_style:
                cell_format[cell.value] = [cell.font, cell.border, cell.fill, cell.number_format, cell.alignment]

        break;

    return cell_format

